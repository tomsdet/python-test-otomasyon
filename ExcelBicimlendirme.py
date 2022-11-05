import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill

dosya = openpyxl.load_workbook("./data.xlsx")
sayfa = dosya["notlar"]
sayfa.cell(row=1, column=4, value="Ortalama")
sayfa.cell(row=1, column=4).font = Font(name="Calibri", sz=16)
satir_sayisi = sayfa.max_row

for satir in range(2,satir_sayisi+1):
    vize = sayfa.cell(satir, 2).value
    final = sayfa.cell(satir, 3).value
    ortalama = (vize + final)/2
    sayfa.cell(row=satir, column=4, value=ortalama)
    sayfa.cell(satir, 4).font = Font(name="Calibri", sz=16, color="2C27CB")
    if ortalama > 70:
        sayfa.cell(satir, 4).fill = PatternFill("solid", "08D84A")
    else:
        sayfa.cell(satir, 4).fill = PatternFill("solid", "E3331B")


# save fonksiyonu verilen konumda verilen isimle dosyanin uzerine yazar
# verilen konumda o isimde dosya yoksa, yeni dosya olusturur
dosya.save("./data.xlsx")