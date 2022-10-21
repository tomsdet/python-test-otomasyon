# dosya  -  sayfa - satir / sutun - hucre - deger
# workbook - sheet - row / column - cell - value

import openpyxl
dosya = openpyxl.load_workbook("./data.xlsx")
print(dosya.sheetnames)
print("aktif sayfa: "+ dosya.active.title)
sayfa = dosya["kitaplar"]
deger = sayfa["B4"].value
print(deger)
#alternatif yontem
deger = dosya["kitaplar"]["B4"].value
print(deger)
#baska bir alternatif
sayfa = dosya["notlar"]
veri = sayfa.cell(2, 2).value
print("veri: "+str(veri))