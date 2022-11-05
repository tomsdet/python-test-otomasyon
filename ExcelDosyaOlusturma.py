import openpyxl
from openpyxl import Workbook

dosya = Workbook()
sayfa = dosya.active
sayfa.title = "hatalar"
sayfa["A1"].value = "python"
sayfa["A2"].value = "excel"
dosya.create_sheet("ikinci", 2)
dosya.create_sheet("ucuncu", 1)

dosya.save("./yeni.xlsx")