import openpyxl

dosya = openpyxl.load_workbook("./data.xlsx")
sayfa = dosya["notlar"]
sayfa.cell(row=1, column=4, value="ortalama")
satir_sayisi = sayfa.max_row

for satir in range(2,satir_sayisi+1):
    vize = sayfa.cell(satir, 2).value
    final = sayfa.cell(satir, 3).value
    ortalama = (vize + final)/2
    sayfa.cell(row=satir, column=4, value=ortalama)

# save fonksiyonu verilen konumda verilen isimle dosyanin uzerine yazar
# verilen konumda o isimde dosya yoksa, yeni dosya olusturur
dosya.save("./data-yeni.xlsx")

